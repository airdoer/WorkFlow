import os
import json
import openpyxl
import io
from Implement.workflowImpl.nodeExecutor import BaseNodeExecutor


class ExcelExecutor(BaseNodeExecutor):
    type = "excel"

    def execute(self, config: dict, input_data: dict) -> dict:
        """
        Excel 渲染节点：接收文件内容或 Table 节点的 tableData 输入。

        输入优先级：
          1. tableData 端口（来自 Table 节点的 tables[0]，直接渲染，跳过文件解析）
          2. fileContent / localPath 端口（来自 P4File / ExcelSearch）

        配置项：
          - sheetName  : 工作表名（可选；input_data['sheetName'] 优先于 config['sheetName']）
          - filterColumns : 列名列表，只保留这些列（空则全部保留）
          - filterRows    : 行号列表（1-based），只保留这些行（空则全部保留）

        输出：
          columns, rows, sheetNames, tableData,
          selectedRows=[], selectedCols=[], selectedValues=null
        """

        # ── 1. tableData 输入（来自 Table 节点 tables 端口） ──────────────────
        # 只在没有 fileContent/localPath 时才走 tableData 路径，
        # 避免 excel 自身上次输出的 tableData 通过端口映射污染本次输入
        local_path_check = input_data.get('localPath', '')
        file_content_check = input_data.get('fileContent', '')
        table_data_input = input_data.get('tableData') if not local_path_check and not file_content_check else None
        if table_data_input:
            # table_data_input 可能是 tables 列表的第一个元素，或 Table 节点直接输出的 tables 列表
            if isinstance(table_data_input, list) and len(table_data_input) > 0:
                first = table_data_input[0]
            elif isinstance(table_data_input, dict):
                first = table_data_input
            else:
                first = None

            if first and isinstance(first, dict) and 'columns' in first and 'rows' in first:
                all_columns = [str(c) for c in first.get('columns', [])]
                raw_rows = first.get('rows', [])
                # allRows_dict: 完整未筛选数据（字典格式）
                all_rows_dict = [{all_columns[i]: (row[i] if isinstance(row, list) and i < len(row) else row.get(all_columns[i]) if isinstance(row, dict) else None) for i in range(len(all_columns))} for row in raw_rows]
                filtered_raw = self._apply_filters_raw(raw_rows, all_columns, config)
                filtered_columns = self._apply_column_filter(all_columns, config)
                rows = [{filtered_columns[i]: (row[i] if isinstance(row, list) and i < len(row) else row.get(filtered_columns[i]) if isinstance(row, dict) else None) for i in range(len(filtered_columns))} for row in filtered_raw]
                table_data = {'title': first.get('title'), 'columns': filtered_columns, 'rows': rows}
                return {
                    'columns': filtered_columns, 'rows': rows,
                    'allColumns': all_columns, 'allRows': all_rows_dict,
                    'sheetNames': [], 'tableData': table_data,
                    'selectedRows': [], 'selectedCols': [], 'selectedValues': None,
                }

        # ── 2. 文件内容输入 ──────────────────────────────────────────────────
        local_path = input_data.get('localPath', '')
        file_content = input_data.get('fileContent', '')

        # sheetName: 连线输入优先于 config
        sheet_name = input_data.get('sheetName') or config.get('sheetName', '') or config.get('sheet', '')

        # filterColumns / filterRows — 支持列表或换行符分隔字符串
        filter_columns = self._parse_list(config.get('filterColumns', ''))
        filter_rows    = self._parse_list(config.get('filterRows',    ''))
        # 兼容旧字段名
        if not filter_columns:
            filter_columns = self._parse_list(config.get('columnFilter', ''))
        if not filter_rows:
            filter_rows = [str(r) for r in (config.get('rowFilter') or [])]

        if not local_path and not file_content:
            return {'error': 'No input. Connect P4File, ExcelSearch, or Table node.'}

        try:
            # 优先用 localPath 打开 xlsx
            if local_path and os.path.exists(local_path):
                wb = openpyxl.load_workbook(local_path, data_only=True)
            elif file_content:
                try:
                    wb = openpyxl.load_workbook(
                        io.BytesIO(
                            file_content.encode('latin-1')
                            if isinstance(file_content, str)
                            else file_content
                        ),
                        data_only=True,
                    )
                except Exception:
                    return self._parse_csv(file_content, filter_rows, filter_columns)
            else:
                return {'error': 'Cannot open file. Ensure an upstream P4 File or ExcelSearch node is connected.'}

            # 选取 sheet
            if sheet_name and str(sheet_name) in wb.sheetnames:
                ws = wb[str(sheet_name)]
            elif sheet_name:
                return {'error': f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"}
            else:
                ws = wb.active

            # 提取原始数据（处理 None 列头）
            raw_columns = [cell.value for cell in ws[1]]
            columns = [str(c) if c is not None else f'Col{i+1}' for i, c in enumerate(raw_columns)]
            all_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                all_rows.append(list(row))

            # 应用筛选
            filtered_rows = self._apply_row_filter(all_rows, filter_rows, columns)
            filtered_columns, filtered_rows = self._apply_column_filter_indexed(columns, filtered_rows, filter_columns)

            rows_dict = [{filtered_columns[i]: (row[i] if i < len(row) else None) for i in range(len(filtered_columns))} for row in filtered_rows]

            table_data = {'title': None, 'columns': filtered_columns, 'rows': rows_dict}
            # allColumns / allRows 保留原始完整数据，供前端筛选 options 使用
            all_rows_dict = [{columns[i]: (row[i] if i < len(row) else None) for i in range(len(columns))} for row in all_rows]
            return {
                'columns': filtered_columns,
                'rows': rows_dict,
                'allColumns': columns,
                'allRows': all_rows_dict,
                'sheetNames': list(wb.sheetnames),
                'tableData': table_data,
                'selectedRows': [],
                'selectedCols': [],
                'selectedValues': None,
            }
        except Exception as e:
            return {'error': str(e)}

    # ── helpers ──────────────────────────────────────────────────────────────

    def _parse_list(self, value) -> list:
        """将列表或换行符/逗号分隔字符串转为列表。"""
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
        if isinstance(value, str) and value.strip():
            # 支持换行符或逗号分隔
            sep = '\n' if '\n' in value else ','
            return [v.strip() for v in value.split(sep) if v.strip()]
        return []

    def _apply_row_filter(self, rows: list, filter_rows: list, columns: list = None) -> list:
        if not filter_rows:
            return rows
        indices = set()
        string_keys = set()
        for r in filter_rows:
            try:
                idx = int(r) - 1  # 1-based → 0-based
                if 0 <= idx < len(rows):
                    indices.add(idx)
                else:
                    # 数字超出范围，尝试作为字符串 key 匹配第一列
                    string_keys.add(r)
            except ValueError:
                string_keys.add(r)

        # 字符串 key → 在第一列中匹配
        if string_keys and columns:
            for i, row in enumerate(rows):
                if i in indices:
                    continue
                first_val = str(row[0]) if len(row) > 0 else ''
                if first_val in string_keys:
                    indices.add(i)

        return [rows[i] for i in sorted(indices)]

    def _apply_column_filter_indexed(self, columns: list, rows: list, filter_columns: list):
        """根据列名过滤列（保留列顺序）。"""
        if not filter_columns:
            return columns, rows
        col_set = set(filter_columns)
        kept_indices = [i for i, c in enumerate(columns) if c in col_set]
        new_columns = [columns[i] for i in kept_indices]
        new_rows = [[row[i] if i < len(row) else None for i in kept_indices] for row in rows]
        return new_columns, new_rows

    def _apply_column_filter(self, columns: list, config: dict) -> list:
        filter_columns = self._parse_list(config.get('filterColumns', ''))
        if not filter_columns:
            return columns
        col_set = set(filter_columns)
        return [c for c in columns if c in col_set]

    def _apply_filters_raw(self, rows: list, columns: list, config: dict) -> list:
        """对 raw rows（列表格式）应用行/列过滤。"""
        filter_rows = self._parse_list(config.get('filterRows', ''))
        rows = self._apply_row_filter(rows, filter_rows, columns)
        return rows

    def _parse_csv(self, content: str, filter_rows: list, filter_columns: list) -> dict:
        """CSV fallback 解析。"""
        import csv
        reader = csv.DictReader(io.StringIO(content))
        columns = list(reader.fieldnames or [])
        rows = list(reader)

        filtered_rows = self._apply_row_filter(rows, filter_rows)

        if filter_columns:
            col_set = set(filter_columns)
            columns = [c for c in columns if c in col_set]
            filtered_rows = [{k: v for k, v in row.items() if k in col_set} for row in filtered_rows]

        table_data = {'title': None, 'columns': columns, 'rows': filtered_rows}
        return {
            'columns': columns,
            'rows': filtered_rows,
            'sheetNames': [],
            'tableData': table_data,
            'selectedRows': [],
            'selectedCols': [],
            'selectedValues': None,
        }
