# -*- coding: utf-8 -*-
# @program c1
# @author chenzhixu@kuaishou.com
# @date 2024/2/7 11:30
# @description:
#   something
import os
from openpyxl import load_workbook
from .XlsxTable import create_xlsx_table, XlsxTable
from .XlsxFile import read_xlsx_directory


g_xlsxCollections = {}


class XlsxCollection(object):
    #  具体svn版本的xlsx集合
    def __init__(self, xlsx_path, svn_version=0):
        self.svn_version = svn_version
        self.xlsx_tables = {}
        self.search_index = {}  # key->list(unique_sheet_name)

        if xlsx_path.endswith('.xlsx'):
            parent_xlsx_path = os.path.basename(xlsx_path)
            self.read_xlsx_file(parent_xlsx_path, xlsx_path)
        self.walk_all_xlsx(xlsx_path)
        self.brief_info = {
            'svn_commit_info': "svn commit info",  # Todo 后续补充上svn的相关信息
            'svn_version': svn_version,
            'sheet_cnt': len(self.xlsx_tables),
            'key_cnt': len(self.search_index),
        }

    def walk_all_xlsx(self, xlsx_directory):
        for file_path in read_xlsx_directory(xlsx_directory):
            self.read_xlsx_file(xlsx_directory, file_path)

    def read_xlsx_file(self, base_directory, file_path):
        # filename = os.path.basename(file_path)
        rel_file_name = os.path.relpath(file_path, base_directory)
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            if rel_file_name.startswith('语言表'):
                continue
            ws = wb[sheet_name]
            unique_sheet_name = f'{rel_file_name}-{sheet_name}'
            print(f'当前处理文件：{rel_file_name}-{sheet_name}')
            meta_info = {
                'svn_version': self.svn_version,
                'file_name': rel_file_name,
                'sheet_name': sheet_name,
                'unique_name': unique_sheet_name,
            }
            xlsx_table: XlsxTable = create_xlsx_table(meta_info, ws)
            if xlsx_table:
                self.add_xlsx_table(unique_sheet_name, xlsx_table)
        wb.close()

    def add_xlsx_table(self, sheet_name: str, xlsx_table: XlsxTable):
        self.xlsx_tables[sheet_name] = xlsx_table
        for k in xlsx_table.search_index.keys():
            if k not in self.search_index:
                self.search_index[k] = []
            if sheet_name not in self.search_index[k]:
                self.search_index[k].append(sheet_name)

    def get_brief_info(self):
        return self.brief_info

    def search_key(self, key: str):
        # Todo 后续需要用前缀树支持前缀搜索，再支持模糊搜索
        key = str(key)  # 全部变成str，因为search_index里都是str的key
        if key not in self.search_index:
            return {}
        else:
            search_result_items = {}  # sheet_name: row_ids
            for sheet_name in self.search_index[key]:
                row_idxs = self.xlsx_tables[sheet_name].search_key(key)
                search_result_items[sheet_name] = row_idxs
        return search_result_items

    def search_keys(self, keys: list[str]):  # 针对多个搜索词
        keys = [str(key) for key in keys]  # 全部变成str，因为search_index里都是str的key

        search_result_items = {}  # sheet_name: row_ids
        valid_sheet_names = set.intersection(*(set(self.search_index.get(k, [])) for k in keys))
        valid_sheet_names = sorted(list(valid_sheet_names))
        for sheet_name in valid_sheet_names:
            row_idxs = self.xlsx_tables[sheet_name].search_keys(keys)
            if not row_idxs:  # 一个表含有搜索的keys，但是不是同一行，返回为空
                continue
            search_result_items[sheet_name] = row_idxs
        return search_result_items

    def get_row_data_by_idxs(self, sheet_name, row_idxs):
        row_datas = []
        for row_idx in row_idxs:
            row_datas.append(self.xlsx_tables[sheet_name].getRowByIdx(row_idx))
        return row_datas

    def get_headers(self, sheet_name):
        return self.xlsx_tables[sheet_name].get_headers()


def initXlsxCollection(svn_version):
    targetXlsxDirectory = f'xlsxImpl/ExcelData/{svn_version}'
    # todo 如果文件不存在，那么就执行svn操作，把对应版本的ExcelDataTable下载下来
    g_xlsxCollections[svn_version] = XlsxCollection(targetXlsxDirectory, svn_version)


def getXlsxCollectionInfo(svn_version):
    if svn_version not in g_xlsxCollections:
        initXlsxCollection(svn_version)
    return g_xlsxCollections[svn_version].get_brief_info()


def requestSearchKey(svn_version: int, search_key: str):
    if svn_version not in g_xlsxCollections:
        return {}
    xlsxCollection = g_xlsxCollections[svn_version]
    search_result = xlsxCollection.search_key(search_key)
    return search_result


def requestSearchKeys(svn_version: int, search_keys: list[str]):
    if svn_version not in g_xlsxCollections:
        return {}
    xlsxCollection = g_xlsxCollections[svn_version]
    search_result = xlsxCollection.search_keys(search_keys)
    return search_result


def requestRowData(svn_version: int, sheet_name: str, row_idxs: list[int]):
    if svn_version not in g_xlsxCollections:
        return []
    xlsxCollection = g_xlsxCollections[svn_version]
    row_datas = xlsxCollection.get_row_data_by_idxs(sheet_name, row_idxs)
    return row_datas


def requestTableHeader(svn_version: int, sheet_name: str):
    if svn_version not in g_xlsxCollections:
        return []
    xlsxCollection = g_xlsxCollections[svn_version]
    headers = xlsxCollection.get_headers(sheet_name)
    return headers


def unitTest():
    xlsxCollection = XlsxCollection("E:/Code/myTools/python/gameXlsx/ExcelDataTable/")
    search_result = xlsxCollection.search_key('204601')
    assert 'character\\角色表.xlsx-角色表' in search_result.keys()
    assert 'combat\\技能表.xlsx-技能' in search_result.keys()
    assert 'combat\\Buff表.xlsx-Buff表' in search_result.keys()
    assert 35 in search_result['character\\角色表.xlsx-角色表']
    assert 167 in search_result['combat\\技能表.xlsx-技能']
    assert 168 in search_result['combat\\技能表.xlsx-技能']
    assert 169 in search_result['combat\\技能表.xlsx-技能']
    assert 1198 in search_result['combat\\Buff表.xlsx-Buff表']

    search_result_for_multi = xlsxCollection.search_keys(['204601'])
    assert 'character\\角色表.xlsx-角色表' in search_result_for_multi.keys()
    assert 'combat\\技能表.xlsx-技能' in search_result_for_multi.keys()
    assert 'combat\\Buff表.xlsx-Buff表' in search_result_for_multi.keys()
    assert 35 in search_result_for_multi['character\\角色表.xlsx-角色表']
    assert 167 in search_result_for_multi['combat\\技能表.xlsx-技能']
    assert 168 in search_result_for_multi['combat\\技能表.xlsx-技能']
    assert 169 in search_result_for_multi['combat\\技能表.xlsx-技能']
    assert 1198 in search_result_for_multi['combat\\Buff表.xlsx-Buff表']

    search_result_for_multi2 = xlsxCollection.search_keys(['204601', '204605'])
    assert 'combat\\技能表.xlsx-技能' in search_result_for_multi2.keys()
    assert 167 in search_result_for_multi2['combat\\技能表.xlsx-技能']
    print(f'search_info_count is {len(xlsxCollection.search_index)}')


if __name__ == "__main__":
    if False:
        t0 = time.time()  # noqa
        # xlsxCollection = XlsxCollection("E:/Code/myTools/python/gameXlsx/ExcelDataTable/")
        # xlsxCollection = XlsxCollection("E:/Project/C1Trunk/Doc/ExcelDataTable/common/关卡任务表.xlsx")
        xlsxCollection = XlsxCollection("E:/Project/C1Trunk/Doc/ExcelDataTable/")
        t1 = time.time()  # noqa
        print(f'XlsxCollection init cost {t1 - t0}, search_info_count is {len(xlsxCollection.search_index)}')

        # 在2024年2月19日的全部导表解析总耗时在21.4s左右
        # 如果增加search_info，总耗时在26s左右，search_index有71139个key，如果不包含语言表目录，search_index有65129个key
    else:
        unitTest()
    a = 1
